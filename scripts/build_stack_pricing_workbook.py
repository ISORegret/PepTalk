"""
Build multi-sheet Excel workbook: grey-market vendor stack pricing, kit vs single,
dosing/reorder, monthly estimates. Run: python scripts/build_stack_pricing_workbook.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Stack-Pricing-Retatrutide-Klow-GH-stack.xlsx"

# Single-vial baseline (user-supplied) — USD
SINGLE = {
    "Retatrutide 10mg": 17.99,
    "Klow 80mg": 40.88,
    "Tesamorelin 10mg": 31.88,
    "Ipamorelin 5mg": 21.44,
    "CJC-1295 (no DAC) 10mg": 29.99,
    "Tesa/Ipa/CJC 6+3+3 blend": 35.22,
}

# Vendor: name, source, Reta10, Klow80, Tesa10, Ipa10, Ipa5, CJC10, CJC5, CP10
# USD per 10-vial kit. None = not on that list.
VENDORS = [
    ("ZLZ (Dec 2025 list)", "PDF: zlz list 12.17 — tier 1–30 boxes", 109, 250, 200, 81, 41, 185, 95, 122),
    ("TCI — Tianjin Cangtu", "PDF: Peptide Price List_TCI — <10 kits", 130, None, 180, 75, 45, 110, 60, 100),
    ("FitPeptideLap warehouse", "PDF: FitPeptideLap updated warehouse pricelist", 120, 260, 210, 150, 80, 140, 80, 150),
    ("Michelle list", "Screenshot (michelleking1015@gmail.com)", 100, 190, 160, 75, None, 140, 80, 135),
    ("Vicky", "Screenshot — 10-vial kits", 88, 170, 148, 65, 45, 158, 75, 100),
    ("BaoHuaDongNuo", "Screenshot", 90, 180, 150, 70, 50, 110, 75, None),
    ("HK peptide / Nicole", "Screenshot — HK USD list; CJC 10 mg N/A on sheet", 109, 240, 175, None, 65, None, 75, 122),
]

FIRST_DATA_ROW = 5
LAST_DATA_ROW = FIRST_DATA_ROW + len(VENDORS) - 1

DEFAULT_MG_WEEK = {
    "Reta": 2.0,
    "Klow": 5.0,
    "Tesa": 2.0,
    "Ipa": 2.0,
    "CJC": 2.0,
}


def style_header(ws, row=1, ncols=20):
    fill = PatternFill("solid", fgColor="1e3a5f")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def main():
    wb = Workbook()

    # --- Sheet 1: Vendor prices ---
    ws1 = wb.active
    ws1.title = "Vendor kit prices"
    ws1["A1"] = (
        "Stack: Retatrutide 10 mg + Klow 80 mg + Tesamorelin 10 mg + Ipamorelin + CJC (no DAC); "
        "optional CP10 (Ipa 5 mg + CJC 5 mg per vial)"
    )
    ws1["A1"].font = Font(bold=True, size=12)
    ws1.merge_cells("A1:J1")
    ws1["A2"] = (
        "Kit = USD for 10 vials. $/vial = kit ÷ 10. Research-use only; verify SKU (no DAC vs DAC), "
        "COA, and shipping. PepTalk PDF was image-only (no extract)."
    )
    ws1.merge_cells("A2:J2")

    headers = [
        "Vendor",
        "Source",
        "Reta 10 mg kit",
        "Klow 80 mg kit",
        "Tesa 10 mg kit",
        "Ipa 10 mg kit",
        "Ipa 5 mg kit",
        "CJC no DAC 10 mg kit",
        "CJC no DAC 5 mg kit",
        "CP10 (5+5) kit",
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=4, column=col, value=h)
    style_header(ws1, row=4, ncols=len(headers))

    for i, row in enumerate(VENDORS, start=FIRST_DATA_ROW):
        for j, val in enumerate(row, start=1):
            ws1.cell(row=i, column=j, value=val)

    stat_row = LAST_DATA_ROW + 2
    ws1.cell(row=stat_row, column=1, value="MIN (best kit $)")
    for col in range(3, 11):
        col_letter = get_column_letter(col)
        ws1.cell(
            row=stat_row,
            column=col,
            value=f"=MIN({col_letter}{FIRST_DATA_ROW}:{col_letter}{LAST_DATA_ROW})",
        )

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 44
    for c in range(3, 12):
        ws1.column_dimensions[get_column_letter(c)].width = 13

    # --- Sheet 2: Best vendor ---
    ws2 = wb.create_sheet("Best vendor picks")
    ws2["A1"] = "Lowest kit $ by column (from Vendor kit prices). Confirm the vendor still lists that SKU."
    ws2.merge_cells("A1:D1")
    ws2["A3"] = "Item"
    ws2["B3"] = "Lowest kit ($)"
    ws2["C3"] = "Vendor"
    style_header(ws2, row=3, ncols=4)

    cols_map = [
        ("Retatrutide 10 mg", "C"),
        ("Klow 80 mg", "D"),
        ("Tesamorelin 10 mg", "E"),
        ("Ipamorelin 10 mg", "F"),
        ("Ipamorelin 5 mg", "G"),
        ("CJC no DAC 10 mg", "H"),
        ("CJC no DAC 5 mg", "I"),
        ("CP10 (5+5)", "J"),
    ]
    r = 4
    for label, col in cols_map:
        lo = FIRST_DATA_ROW
        hi = LAST_DATA_ROW
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=f"=MIN('Vendor kit prices'!{col}{lo}:{col}{hi})")
        ws2.cell(
            row=r,
            column=3,
            value=(
                f"=INDEX('Vendor kit prices'!$A${lo}:$A${hi},"
                f"MATCH(MIN('Vendor kit prices'!{col}{lo}:{col}{hi}),"
                f"'Vendor kit prices'!{col}{lo}:{col}{hi},0))"
            ),
        )
        r += 1

    ws2["A14"] = "Notes"
    ws2["A15"] = (
        "• TCI Klow 80 mg: not on extracted PDF pages — left blank.\n"
        "• HK peptide: no 10 mg CJC price on your screenshot; 5 mg kit shown.\n"
        "• 6/3/3 Tesa/Ipa/CJC blend is not CP10; price single vial on Kit vs single sheet only.\n"
        "• Cheapest per component may be split across vendors — add shipping before deciding."
    )
    ws2["A15"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A15:D19")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 30

    # --- Sheet 3: Kit vs single ---
    ws3 = wb.create_sheet("Kit vs single vial")
    ws3["A1"] = "10 × your single-vial price vs lowest 10-vial kit from the table (same nominal strength per vial)."
    ws3.merge_cells("A1:F1")
    hdr = ["Product", "Single-vial ($)", "10 × singles ($)", "MIN kit ($)", "$/vial from kit", "Savings (10 singles − kit)"]
    for c, h in enumerate(hdr, 1):
        ws3.cell(row=3, column=c, value=h)
    style_header(ws3, row=3, ncols=6)

    kit_min = {
        "Reta": "=MIN('Vendor kit prices'!C5:C11)",
        "Klow": "=MIN('Vendor kit prices'!D5:D11)",
        "Tesa": "=MIN('Vendor kit prices'!E5:E11)",
        "Ipa5": "=MIN('Vendor kit prices'!G5:G11)",
        "CJC10": "=MIN('Vendor kit prices'!H5:H11)",
        "CP10": "=MIN('Vendor kit prices'!J5:J11)",
    }

    rows_ks = [
        ("Retatrutide 10 mg", SINGLE["Retatrutide 10mg"], kit_min["Reta"]),
        ("Klow 80 mg", SINGLE["Klow 80mg"], kit_min["Klow"]),
        ("Tesamorelin 10 mg", SINGLE["Tesamorelin 10mg"], kit_min["Tesa"]),
        ("Ipamorelin 5 mg single (your price)", SINGLE["Ipamorelin 5mg"], kit_min["Ipa5"]),
        ("CJC no DAC 10 mg", SINGLE["CJC-1295 (no DAC) 10mg"], kit_min["CJC10"]),
        ("CP10 (5+5) kit — no single-vial baseline", None, kit_min["CP10"]),
        ("Tesa/Ipa/CJC 6+3+3 blend (your single)", SINGLE["Tesa/Ipa/CJC 6+3+3 blend"], "—"),
    ]

    for i, (name, single, kit_formula) in enumerate(rows_ks, start=4):
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=single)
        ws3.cell(row=i, column=3, value=f"=IF(B{i}=\"\",\"\",B{i}*10)")
        ws3.cell(row=i, column=4, value=kit_formula)
        ws3.cell(
            row=i,
            column=5,
            value=f"=IF(OR(D{i}=\"\",D{i}=\"—\"),\"\",D{i}/10)",
        )
        ws3.cell(
            row=i,
            column=6,
            value=f"=IF(OR(D{i}=\"\",D{i}=\"—\"),\"\",C{i}-D{i})",
        )

    r_note = 4 + len(rows_ks) + 1
    ws3.cell(row=r_note, column=1, value="CP10 vs separate 5 mg kits (same total mg per vial)")
    ws3.cell(
        row=r_note,
        column=4,
        value=(
            "=MIN('Vendor kit prices'!J5:J11)-"
            "MIN('Vendor kit prices'!G5:G11)-MIN('Vendor kit prices'!I5:I11)"
        ),
    )
    ws3.cell(row=r_note, column=5, value="Negative = CP10 cheaper than Ipa5 + CJC5 kits")
    ws3.merge_cells(f"E{r_note}:F{r_note}")

    ws3.cell(row=r_note + 2, column=1, value=(
        "Longevity: a 10-vial kit lasts 10× as many weeks as one vial at the same mg/week. "
        "If singles cost less per vial than kit÷10, kits still save only if you will use all 10 vials."
    ))
    ws3.cell(row=r_note + 2, column=1).alignment = Alignment(wrap_text=True)
    ws3.merge_cells(f"A{r_note + 2}:F{r_note + 3}")
    ws3.column_dimensions["A"].width = 46
    for c in "BCDEF":
        ws3.column_dimensions[c].width = 16

    # --- Sheet 4: Dosing ---
    ws4 = wb.create_sheet("Dosing and reorder")
    ws4["A1"] = "Edit mg/week (column C) to match your protocol. Defaults are placeholders — not medical advice."
    ws4.merge_cells("A1:E1")
    h2 = ["Line", "Vial total (mg)", "mg per week (edit)", "Weeks per vial", "Weeks per 10-vial kit"]
    for c, h in enumerate(h2, 1):
        ws4.cell(row=3, column=c, value=h)
    style_header(ws4, row=3, ncols=5)

    dosing = [
        ("Retatrutide 10 mg", 10, DEFAULT_MG_WEEK["Reta"]),
        ("Klow 80 mg (blend total)", 80, DEFAULT_MG_WEEK["Klow"]),
        ("Tesamorelin 10 mg", 10, DEFAULT_MG_WEEK["Tesa"]),
        ("Ipamorelin 10 mg", 10, DEFAULT_MG_WEEK["Ipa"]),
        ("CJC no DAC 10 mg", 10, DEFAULT_MG_WEEK["CJC"]),
    ]
    for i, (name, vial_mg, mg_w) in enumerate(dosing, start=4):
        ws4.cell(row=i, column=1, value=name)
        ws4.cell(row=i, column=2, value=vial_mg)
        ws4.cell(row=i, column=3, value=mg_w)
        ws4.cell(row=i, column=4, value=f"=IF(C{i}=0,\"\",B{i}/C{i})")
        ws4.cell(row=i, column=5, value=f"=IF(C{i}=0,\"\",D{i}*10)")

    ws4["A11"] = "Reorder: set a reminder when remaining weeks ≤ your shipping lead time (e.g. 2 weeks)."
    ws4.merge_cells("A11:E11")

    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["C"].width = 18

    # --- Sheet 5: Monthly cost ---
    ws5 = wb.create_sheet("Monthly stack cost")
    ws5["A1"] = (
        "$/month ≈ ($/vial ÷ weeks per vial) × 4.345. Uses MIN kit $ and Dosing sheet weeks. "
        "Split vendors = sum still valid for budget."
    )
    ws5.merge_cells("A1:E1")
    wh = ["Component", "MIN kit ($)", "$/vial", "Weeks/vial (Dosing)", "$/month"]
    for c, h in enumerate(wh, 1):
        ws5.cell(row=3, column=c, value=h)
    style_header(ws5, row=3, ncols=5)

    # Dosing rows: Reta=4, Klow=5, Tesa=6, Ipa=7, CJC=8
    monthly = [
        ("Retatrutide 10 mg", "C", "Dosing and reorder!D4"),
        ("Klow 80 mg", "D", "Dosing and reorder!D5"),
        ("Tesamorelin 10 mg", "E", "Dosing and reorder!D6"),
        ("Ipamorelin 10 mg", "F", "Dosing and reorder!D7"),
        ("CJC no DAC 10 mg", "H", "Dosing and reorder!D8"),
    ]
    r = 4
    for name, col, weeks_ref in monthly:
        ws5.cell(row=r, column=1, value=name)
        ws5.cell(
            row=r,
            column=2,
            value=f"=MIN('Vendor kit prices'!{col}{FIRST_DATA_ROW}:{col}{LAST_DATA_ROW})",
        )
        ws5.cell(row=r, column=3, value=f"=B{r}/10")
        ws5.cell(row=r, column=4, value=f"={weeks_ref}")
        ws5.cell(row=r, column=5, value=f"=IF(D{r}=0,\"\",(C{r}/D{r})*4.345)")
        r += 1

    ws5.cell(row=9, column=1, value="Approx. stack total / month")
    ws5.cell(row=9, column=5, value="=SUM(E4:E8)")

    ws5.column_dimensions["A"].width = 22
    for c in "BCDE":
        ws5.column_dimensions[c].width = 14

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
